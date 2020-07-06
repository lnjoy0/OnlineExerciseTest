from django.contrib import admin
from .models import *
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from django.utils.html import format_html
from django.utils import timezone

# Register your models here.
admin.site.site_header='在线习题测试系统后台'
admin.site.site_title='在线习题测试系统'

class SCInLine(admin.TabularInline):
    model = SC_question
    extra = 2

class MCInLine(admin.TabularInline):
    model = MC_question
    extra = 2

class BlankInLine(admin.TabularInline):
    model = Blank_question
    extra = 2

class PaperAdmin(admin.ModelAdmin):
    #inlines = [SCInLine]+[MCInLine]+[BlankInLine]
    list_display = ('paper_text', 'pub_date')
    list_display_links = ('paper_text','pub_date')
    search_fields = ['paper_text']
    list_filter = ['pub_date']

class ScoreAdmin(admin.ModelAdmin):
    list_display = ('score', 'user_id', 'paper_id')
    list_display_links = ('score', 'user_id', 'paper_id')
    exclude = ('answer',)

class SC_choiceInLine(admin.TabularInline):
    model = SC_choice
    extra = 4

class SC_questionAdmin(admin.ModelAdmin):
    list_display = ('id', 'SC_text', 'value', 'paper')
    list_display_links = ('id', 'SC_text', 'value', 'paper')
    inlines = [SC_choiceInLine]

class MC_choiceInLine(admin.TabularInline):
    model = MC_choice
    extra = 4

class MC_questionAdmin(admin.ModelAdmin):
    list_display = ('id', 'MC_text', 'value', 'paper')
    list_display_links = ('id', 'MC_text', 'value', 'paper')
    inlines = [MC_choiceInLine]

class Blank_questionAdmin(admin.ModelAdmin):
    list_display = ('id', 'blank_text', 'value', 'paper')
    list_display_links = ('id', 'blank_text', 'value', 'paper')

class FileAdmin(admin.ModelAdmin):
    actions = ['download',]

    def download(self,request,queryset):
        pass

    download.type = 'success'
    download.action_type = 1
    download.action_url = '/download'
    download.short_description = '下载试题模板'

    list_display = ('file','name')
    list_filter = ['name',]
    def save_model(self, request, obj, form, change):
        re = super(FileAdmin,self).save_model(request,obj,form,change)
        wb = load_workbook(filename=obj.file.path)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        headers = ['问题描述','题型','正确答案','分值','选项A','选项B','选项C','选项D','选项E','选项F']
        lists = []
        for row in range(3,300):
            if ws.cell(row=row, column=1).value == None:
                break
            r = {}
            for col in range(1,len(headers)):
                if ws.cell(row=row, column=col).value == None:
                    break
                key = headers[col - 1]
                r[key] = ws.cell(row=row, column=col).value
            lists.append(r)
        new_paper = Paper(paper_text=obj.name, pub_date=timezone.now(), time=obj.time)
        new_paper.save()
        for i in lists:
            if i['题型'] == '单选题':
                new_sc_question = SC_question(SC_text=i['问题描述'], SC_solution=i['正确答案'], value=i['分值'], paper=new_paper)
                new_sc_question.save()
                for j in list('ABCDEF'):
                    if '选项'+j in i.keys():
                        new_sc_choice = SC_choice(choice_symbol=j, choice_text=i['选项'+j], question=new_sc_question, paper=new_paper)
                        new_sc_choice.save()
            elif i['题型'] == '多选题':
                new_mc_question = MC_question(MC_text=i['问题描述'], MC_solution=i['正确答案'], value=i['分值'], paper=new_paper)
                new_mc_question.save()
                for j in list('ABCDEF'):
                    if '选项'+j in i.keys():
                        new_mc_choice = MC_choice(choice_symbol=j, choice_text=i['选项'+j], question=new_mc_question, paper=new_paper)
                        new_mc_choice.save()
            else:
                new_blank_question = Blank_question(blank_text=i['问题描述'], blank_solution=i['正确答案'], value=i['分值'], paper=new_paper)
                new_blank_question.save()

admin.site.register(Paper, PaperAdmin)
#admin.site.register(SC_choice)
#admin.site.register(MC_choice)
admin.site.register(Score, ScoreAdmin)
admin.site.register(User)
admin.site.register(SC_question,SC_questionAdmin)
admin.site.register(MC_question,MC_questionAdmin)
admin.site.register(Blank_question,Blank_questionAdmin)
admin.site.register(File, FileAdmin)
